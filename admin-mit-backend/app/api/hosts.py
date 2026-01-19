"""
SSH 主机管理 API
"""
from flask import Blueprint, request, jsonify, g
from app.services.ssh_service import ssh_service
from app.services.host_info_service import host_info_service
from app.models.host import SSHHost, HostInfo
from app.extensions import db
from app.core.middleware import tenant_required, role_required
from app.services.password_service import password_decrypt_service
from datetime import datetime
import logging

logger = logging.getLogger(__name__)
hosts_bp = Blueprint('hosts', __name__)


@hosts_bp.route('', methods=['GET'])
@tenant_required
def get_hosts():
    """获取主机列表"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', '')
        group_id = request.args.get('group_id', type=int)
        sort_by = request.args.get('sort_by', 'name')  # 排序字段
        sort_order = request.args.get('sort_order', 'asc')  # 排序方向: asc/desc
        
        # 限制每页最大数量
        per_page = min(per_page, 100)
        
        # 构建查询
        query = SSHHost.query.filter_by(tenant_id=g.tenant_id)
        
        # 分组筛选功能
        if group_id is not None:
            if group_id == 0:
                # group_id=0 表示筛选未分组的主机
                query = query.filter(SSHHost.group_id.is_(None))
            else:
                query = query.filter(SSHHost.group_id == group_id)
        
        # 搜索功能
        if search:
            search_pattern = f'%{search}%'
            query = query.filter(
                db.or_(
                    SSHHost.name.ilike(search_pattern),
                    SSHHost.hostname.ilike(search_pattern),
                    SSHHost.description.ilike(search_pattern)
                )
            )
        
        # 排序功能
        sort_columns = {
            'name': SSHHost.name,
            'hostname': SSHHost.hostname,
            'status': SSHHost.last_probe_status,
            'probe_status': SSHHost.last_probe_status,
            'last_probe_at': SSHHost.last_probe_at,
            'created_at': SSHHost.created_at,
            'updated_at': SSHHost.updated_at,
        }
        
        sort_column = sort_columns.get(sort_by, SSHHost.name)
        if sort_order == 'desc':
            query = query.order_by(sort_column.desc().nulls_last())
        else:
            query = query.order_by(sort_column.asc().nulls_last())
        
        # 分页
        pagination = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        hosts = []
        for host in pagination.items:
            host_data = host.to_dict(include_sensitive=False)
            # 添加最新性能指标
            latest_metrics = host_info_service.get_host_latest_metrics(host.id)
            host_data['latest_metrics'] = latest_metrics
            hosts.append(host_data)
        
        return jsonify({
            'success': True,
            'data': {
                'hosts': hosts,
                'pagination': {
                    'page': pagination.page,
                    'per_page': pagination.per_page,
                    'total': pagination.total,
                    'pages': pagination.pages,
                    'has_prev': pagination.has_prev,
                    'has_next': pagination.has_next
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Get hosts error: {e}")
        return jsonify({
            'success': False,
            'message': '获取主机列表失败'
        }), 500


@hosts_bp.route('', methods=['POST'])
@role_required('admin', 'super_admin', '超级管理员', '运维管理员', '系统管理员')
def create_host():
    """创建主机"""
    try:
        data = request.get_json()
        
        # 验证必需字段
        required_fields = ['name', 'hostname', 'username', 'auth_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'缺少必需字段: {field}'
                }), 400
        
        # 验证认证类型
        auth_type = data.get('auth_type')
        if auth_type not in ['password', 'key']:
            return jsonify({
                'success': False,
                'message': '认证类型必须是 password 或 key'
            }), 400
        
        # 验证认证信息
        if auth_type == 'password' and not data.get('password'):
            return jsonify({
                'success': False,
                'message': '密码认证需要提供密码'
            }), 400
        
        if auth_type == 'key' and not data.get('private_key'):
            return jsonify({
                'success': False,
                'message': '密钥认证需要提供私钥'
            }), 400
        
        # 检查主机名是否已存在
        existing_host = SSHHost.query.filter_by(
            tenant_id=g.tenant_id,
            hostname=data['hostname'],
            port=data.get('port', 22)
        ).first()
        
        if existing_host:
            return jsonify({
                'success': False,
                'message': '该主机地址已存在'
            }), 400
        
        # 处理密码：前端加密 → 后端解密后存储明文
        decrypted_password = None
        if auth_type == 'password' and data.get('password'):
            try:
                decrypted_password = password_decrypt_service.decrypt_password(data['password'], data['name'], data['hostname'])
            except Exception as e:
                logger.warning(f"密码解密失败，使用原始值: {e}")
                decrypted_password = data['password']
        
        # 创建主机记录
        host = SSHHost(
            tenant_id=g.tenant_id,
            name=data['name'],
            hostname=data['hostname'],
            port=data.get('port', 22),
            username=data['username'],
            auth_type=auth_type,
            password=decrypted_password,  # 存储解密后的明文密码
            private_key=data.get('private_key'),
            description=data.get('description', ''),
            os_type=data.get('os_type', ''),
            group_id=data.get('group_id'),
            status=1
        )
        
        db.session.add(host)
        db.session.commit()
        
        # 测试连接
        try:
            success, message = ssh_service.test_connection(
                hostname=host.hostname,
                port=host.port,
                username=host.username,
                auth_type=host.auth_type,
                password=host.password,
                private_key=host.private_key
            )
            
            if success:
                # 连接成功，收集系统信息
                try:
                    host_info_service.collect_host_system_info(host)
                    logger.info(f"成功收集主机系统信息: {host.name}")
                except Exception as e:
                    logger.warning(f"收集主机系统信息失败: {e}")
            else:
                logger.warning(f"主机连接测试失败: {message}")
                
        except Exception as e:
            logger.warning(f"主机连接测试异常: {e}")
        
        return jsonify({
            'success': True,
            'data': {
                'host': host.to_dict(include_sensitive=False)
            },
            'message': '主机创建成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Create host error: {e}")
        return jsonify({
            'success': False,
            'message': '创建主机失败'
        }), 500


@hosts_bp.route('/<int:host_id>', methods=['PUT'])
@role_required('admin', 'super_admin', '超级管理员', '运维管理员', '系统管理员')
def update_host(host_id):
    """更新主机"""
    try:
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        data = request.get_json()
        
        # 更新基本信息
        if 'name' in data:
            host.name = data['name']
        if 'hostname' in data:
            host.hostname = data['hostname']
        if 'port' in data:
            host.port = data['port']
        if 'username' in data:
            host.username = data['username']
        if 'description' in data:
            host.description = data['description']
        if 'os_type' in data:
            host.os_type = data['os_type']
        if 'status' in data:
            host.status = data['status']
        if 'group_id' in data:
            # 支持设置为 None（未分组）或有效的分组 ID
            host.group_id = data['group_id'] if data['group_id'] else None
        
        # 更新认证信息
        if 'auth_type' in data:
            auth_type = data['auth_type']
            if auth_type not in ['password', 'key']:
                return jsonify({
                    'success': False,
                    'message': '认证类型必须是 password 或 key'
                }), 400
            
            host.auth_type = auth_type
            
            # 清除旧的认证信息
            host.password = None
            host.private_key = None
            
            # 设置新的认证信息
            if auth_type == 'password':
                if not data.get('password'):
                    return jsonify({
                        'success': False,
                        'message': '密码认证需要提供密码'
                    }), 400
                # 解密前端加密的密码后存储
                try:
                    host.password = password_decrypt_service.decrypt_password(data['password'], host.name, host.hostname)
                except Exception as e:
                    logger.warning(f"密码解密失败，使用原始值: {e}")
                    host.password = data['password']
            elif auth_type == 'key':
                if not data.get('private_key'):
                    return jsonify({
                        'success': False,
                        'message': '密钥认证需要提供私钥'
                    }), 400
                host.private_key = data['private_key']
        
        # 如果只更新密码或私钥
        if 'password' in data and host.auth_type == 'password':
            # 解密前端加密的密码后存储
            try:
                host.password = password_decrypt_service.decrypt_password(data['password'], host.name, host.hostname)
            except Exception as e:
                logger.warning(f"密码解密失败，使用原始值: {e}")
                host.password = data['password']
        if 'private_key' in data and host.auth_type == 'key':
            host.private_key = data['private_key']
        
        db.session.commit()
        
        # 如果更新了连接信息，测试连接
        connection_fields = ['hostname', 'port', 'username', 'auth_type', 'password', 'private_key']
        if any(field in data for field in connection_fields):
            try:
                success, message = ssh_service.test_connection(
                    hostname=host.hostname,
                    port=host.port,
                    username=host.username,
                    auth_type=host.auth_type,
                    password=host.password,
                    private_key=host.private_key
                )
                
                if success:
                    # 连接成功，更新系统信息
                    try:
                        host_info_service.collect_host_system_info(host)
                        logger.info(f"成功更新主机系统信息: {host.name}")
                    except Exception as e:
                        logger.warning(f"更新主机系统信息失败: {e}")
                else:
                    logger.warning(f"主机连接测试失败: {message}")
                    
            except Exception as e:
                logger.warning(f"主机连接测试异常: {e}")
        
        return jsonify({
            'success': True,
            'data': {
                'host': host.to_dict(include_sensitive=False)
            },
            'message': '主机更新成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Update host error: {e}")
        return jsonify({
            'success': False,
            'message': '更新主机失败'
        }), 500


@hosts_bp.route('/<int:host_id>/delete', methods=['POST'])
@role_required('admin', 'super_admin', '超级管理员', '运维管理员', '系统管理员')
def delete_host(host_id):
    """删除主机"""
    try:
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        # 停止该主机的性能数据采集
        try:
            host_info_service.stop_periodic_collection(host)
        except Exception as e:
            logger.warning(f"停止主机性能采集失败: {e}")
        
        # 删除主机（级联删除相关数据）
        db.session.delete(host)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '主机删除成功'
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Delete host error: {e}")
        return jsonify({
            'success': False,
            'message': '删除主机失败'
        }), 500


@hosts_bp.route('/<int:host_id>', methods=['GET'])
@tenant_required
def get_host(host_id):
    """获取单个主机详情"""
    try:
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        host_data = host.to_dict(include_sensitive=False)
        
        # 添加最新性能指标
        latest_metrics = host_info_service.get_host_latest_metrics(host.id)
        host_data['latest_metrics'] = latest_metrics
        
        # 添加性能指标历史（最近24小时）
        metrics_history = host_info_service.get_host_metrics_history(host.id, 24)
        host_data['metrics_history'] = metrics_history
        
        return jsonify({
            'success': True,
            'data': {
                'host': host_data
            }
        })
        
    except Exception as e:
        logger.error(f"Get host error: {e}")
        return jsonify({
            'success': False,
            'message': '获取主机详情失败'
        }), 500


@hosts_bp.route('/<int:host_id>/status', methods=['GET'])
@tenant_required
def get_host_status(host_id):
    """获取主机状态"""
    try:
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        # 获取最新性能指标
        latest_metrics = host_info_service.get_host_latest_metrics(host.id)
        
        # 获取采集状态
        collection_status = host_info_service.get_collection_status()
        is_collecting = f"host_{host.id}" in collection_status.get('collecting_hosts', [])
        
        return jsonify({
            'success': True,
            'data': {
                'host_id': host.id,
                'name': host.name,
                'hostname': host.hostname,
                'status': host.status,
                'last_connected_at': host.last_connected_at.isoformat() if host.last_connected_at else None,
                'latest_metrics': latest_metrics,
                'is_collecting': is_collecting
            }
        })
        
    except Exception as e:
        logger.error(f"Get host status error: {e}")
        return jsonify({
            'success': False,
            'message': '获取主机状态失败'
        }), 500


@hosts_bp.route('/<int:host_id>/connect', methods=['POST'])
@tenant_required
def test_host_connection(host_id):
    """测试主机连接"""
    try:
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        # 测试连接
        success, message = ssh_service.test_connection(
            hostname=host.hostname,
            port=host.port,
            username=host.username,
            auth_type=host.auth_type,
            password=host.password,
            private_key=host.private_key
        )
        
        # 更新最后连接时间
        if success:
            host.last_connected_at = datetime.utcnow()
            db.session.commit()
        
        return jsonify({
            'success': success,
            'data': {
                'host_id': host.id,
                'name': host.name,
                'hostname': host.hostname,
                'connected': success,
                'message': message,
                'tested_at': datetime.utcnow().isoformat()
            },
            'message': message
        })
        
    except Exception as e:
        logger.error(f"Test host connection error: {e}")
        return jsonify({
            'success': False,
            'message': '测试主机连接失败'
        }), 500


@hosts_bp.route('/<int:host_id>/info', methods=['GET'])
@tenant_required
def get_host_info(host_id):
    """获取主机系统信息"""
    try:
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        # 获取主机系统信息
        host_info = HostInfo.query.filter_by(host_id=host.id).first()
        
        if not host_info:
            # 如果没有系统信息，尝试收集
            try:
                system_info = host_info_service.collect_host_system_info(host)
                host_info = HostInfo.query.filter_by(host_id=host.id).first()
            except Exception as e:
                logger.warning(f"收集主机系统信息失败: {e}")
                return jsonify({
                    'success': False,
                    'message': '无法获取主机系统信息，请检查主机连接'
                }), 500
        
        # 构建返回数据
        info_data = {
            'host_id': host.id,
            'name': host.name,
            'hostname': host.hostname,
            'port': host.port,
            'username': host.username,
            'auth_type': host.auth_type,
            'description': host.description,
            'status': host.status,
            'last_connected_at': host.last_connected_at.isoformat() if host.last_connected_at else None,
            'created_at': host.created_at.isoformat(),
            'updated_at': host.updated_at.isoformat()
        }
        
        if host_info:
            info_data.update({
                'system_info': {
                    'os_name': host_info.os_name,
                    'os_version': host_info.os_version,
                    'kernel_version': host_info.kernel_version,
                    'cpu_cores': host_info.cpu_cores,
                    'total_memory': host_info.total_memory,
                    'disk_total': host_info.disk_total,
                    'network_interfaces': host_info.network_interfaces,
                    'updated_at': host_info.updated_at.isoformat()
                }
            })
        
        return jsonify({
            'success': True,
            'data': info_data
        })
        
    except Exception as e:
        logger.error(f"Get host info error: {e}")
        return jsonify({
            'success': False,
            'message': '获取主机信息失败'
        }), 500


@hosts_bp.route('/<int:host_id>/metrics', methods=['GET'])
@tenant_required
def get_host_metrics(host_id):
    """获取主机性能监控数据"""
    try:
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        # 获取查询参数
        hours = request.args.get('hours', 24, type=int)
        limit = request.args.get('limit', 100, type=int)
        
        # 限制查询范围
        hours = min(hours, 168)  # 最多7天
        limit = min(limit, 1000)  # 最多1000条记录
        
        # 获取最新性能指标
        latest_metrics = host_info_service.get_host_latest_metrics(host.id)
        
        # 获取历史性能指标
        metrics_history = host_info_service.get_host_metrics_history(host.id, hours)
        
        # 如果需要限制记录数，取最新的记录
        if len(metrics_history) > limit:
            metrics_history = metrics_history[-limit:]
        
        # 获取采集状态
        collection_status = host_info_service.get_collection_status()
        is_collecting = f"host_{host.id}" in collection_status.get('collecting_hosts', [])
        
        # 计算统计信息
        stats = {}
        if metrics_history:
            cpu_values = [m['cpu_usage'] for m in metrics_history if m.get('cpu_usage') is not None]
            memory_values = [m['memory_usage'] for m in metrics_history if m.get('memory_usage') is not None]
            disk_values = [m['disk_usage'] for m in metrics_history if m.get('disk_usage') is not None]
            
            if cpu_values:
                stats['cpu'] = {
                    'avg': round(sum(cpu_values) / len(cpu_values), 2),
                    'max': max(cpu_values),
                    'min': min(cpu_values)
                }
            
            if memory_values:
                stats['memory'] = {
                    'avg': round(sum(memory_values) / len(memory_values), 2),
                    'max': max(memory_values),
                    'min': min(memory_values)
                }
            
            if disk_values:
                stats['disk'] = {
                    'avg': round(sum(disk_values) / len(disk_values), 2),
                    'max': max(disk_values),
                    'min': min(disk_values)
                }
        
        return jsonify({
            'success': True,
            'data': {
                'host_id': host.id,
                'name': host.name,
                'hostname': host.hostname,
                'latest_metrics': latest_metrics,
                'metrics_history': metrics_history,
                'statistics': stats,
                'is_collecting': is_collecting,
                'query_params': {
                    'hours': hours,
                    'limit': limit,
                    'total_records': len(metrics_history)
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Get host metrics error: {e}")
        return jsonify({
            'success': False,
            'message': '获取主机性能监控数据失败'
        }), 500


@hosts_bp.route('/<int:host_id>/webshell', methods=['POST'])
@tenant_required
def create_webshell_session(host_id):
    """创建 WebShell 会话"""
    try:
        from app.services.webshell_service import webshell_service
        from app.services.webshell_terminal_service import webshell_terminal_service
        
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        data = request.get_json() or {}
        cols = data.get('cols', 80)
        rows = data.get('rows', 24)
        
        # 创建 WebShell 会话
        success, message, session_data = webshell_service.create_session(
            host_id=host.id,
            user_id=g.user_id,
            tenant_id=g.tenant_id
        )
        
        if not success:
            return jsonify({
                'success': False,
                'message': message
            }), 400
        
        session_id = session_data['session_id']
        
        # 创建终端会话
        terminal_success, terminal_message = webshell_terminal_service.create_terminal_session(
            session_id, cols, rows
        )
        
        if not terminal_success:
            # 如果终端会话创建失败，清理 WebShell 会话
            webshell_service.terminate_session(session_id)
            return jsonify({
                'success': False,
                'message': f'创建终端会话失败: {terminal_message}'
            }), 500
        
        return jsonify({
            'success': True,
            'data': {
                'session_id': session_id,
                'host_id': host.id,
                'host_name': host.name,
                'hostname': host.hostname,
                'username': host.username,
                'terminal_size': {
                    'cols': cols,
                    'rows': rows
                },
                'created_at': datetime.utcnow().isoformat()
            },
            'message': 'WebShell 会话创建成功'
        })
        
    except Exception as e:
        logger.error(f"Create WebShell session error: {e}")
        return jsonify({
            'success': False,
            'message': '创建 WebShell 会话失败'
        }), 500


@hosts_bp.route('/webshell/<session_id>', methods=['GET'])
@tenant_required
def get_webshell_session(session_id):
    """获取 WebShell 会话信息"""
    try:
        from app.services.webshell_service import webshell_service
        
        session_info = webshell_service.get_session(session_id)
        
        if not session_info:
            return jsonify({
                'success': False,
                'message': 'WebShell 会话不存在'
            }), 404
        
        # 验证会话所有权
        if session_info['user_id'] != g.user_id:
            return jsonify({
                'success': False,
                'message': '无权访问该会话'
            }), 403
        
        # 获取主机信息
        host = SSHHost.query.filter_by(
            id=session_info['host_id'],
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '关联的主机不存在'
            }), 404
        
        return jsonify({
            'success': True,
            'data': {
                'session_id': session_id,
                'host_id': host.id,
                'host_name': host.name,
                'hostname': host.hostname,
                'username': host.username,
                'status': session_info['status'],
                'terminal_size': {
                    'cols': session_info.get('cols', 80),
                    'rows': session_info.get('rows', 24)
                },
                'created_at': session_info['created_at'],
                'last_activity': session_info.get('last_activity'),
                'is_active': session_info['status'] == 'active'
            }
        })
        
    except Exception as e:
        logger.error(f"Get WebShell session error: {e}")
        return jsonify({
            'success': False,
            'message': '获取 WebShell 会话信息失败'
        }), 500


@hosts_bp.route('/webshell/<session_id>/status', methods=['GET'])
@tenant_required
def get_webshell_session_status(session_id):
    """获取 WebShell 会话状态"""
    try:
        from app.services.webshell_service import webshell_service
        from app.services.webshell_terminal_service import webshell_terminal_service
        
        session_info = webshell_service.get_session(session_id)
        
        if not session_info:
            return jsonify({
                'success': False,
                'message': 'WebShell 会话不存在'
            }), 404
        
        # 验证会话所有权
        if session_info['user_id'] != g.user_id:
            return jsonify({
                'success': False,
                'message': '无权访问该会话'
            }), 403
        
        # 获取终端会话状态
        terminal_session = webshell_terminal_service.get_terminal_session(session_id)
        terminal_active = terminal_session.is_active if terminal_session else False
        
        return jsonify({
            'success': True,
            'data': {
                'session_id': session_id,
                'webshell_status': session_info['status'],
                'terminal_active': terminal_active,
                'last_activity': session_info.get('last_activity'),
                'created_at': session_info['created_at'],
                'is_connected': session_info['status'] == 'active' and terminal_active
            }
        })
        
    except Exception as e:
        logger.error(f"Get WebShell session status error: {e}")
        return jsonify({
            'success': False,
            'message': '获取 WebShell 会话状态失败'
        }), 500


@hosts_bp.route('/webshell/<session_id>/history', methods=['GET'])
@tenant_required
def get_webshell_command_history(session_id):
    """获取 WebShell 命令历史"""
    try:
        from app.services.webshell_service import webshell_service
        from app.services.webshell_terminal_service import webshell_terminal_service
        
        session_info = webshell_service.get_session(session_id)
        
        if not session_info:
            return jsonify({
                'success': False,
                'message': 'WebShell 会话不存在'
            }), 404
        
        # 验证会话所有权
        if session_info['user_id'] != g.user_id:
            return jsonify({
                'success': False,
                'message': '无权访问该会话'
            }), 403
        
        # 获取查询参数
        limit = request.args.get('limit', 100, type=int)
        limit = min(limit, 1000)  # 最多1000条记录
        
        # 获取命令历史
        history = webshell_terminal_service.get_command_history(session_id, limit)
        
        return jsonify({
            'success': True,
            'data': {
                'session_id': session_id,
                'history': history,
                'total_count': len(history)
            }
        })
        
    except Exception as e:
        logger.error(f"Get WebShell command history error: {e}")
        return jsonify({
            'success': False,
            'message': '获取命令历史失败'
        }), 500


@hosts_bp.route('/webshell/<session_id>/execute', methods=['POST'])
@tenant_required
def execute_webshell_command(session_id):
    """执行 WebShell 命令"""
    try:
        from app.services.webshell_service import webshell_service
        from app.services.webshell_terminal_service import webshell_terminal_service
        
        session_info = webshell_service.get_session(session_id)
        
        if not session_info:
            return jsonify({
                'success': False,
                'message': 'WebShell 会话不存在'
            }), 404
        
        # 验证会话所有权
        if session_info['user_id'] != g.user_id:
            return jsonify({
                'success': False,
                'message': '无权访问该会话'
            }), 403
        
        data = request.get_json()
        command = data.get('command', '').strip()
        timeout = data.get('timeout', 30)
        
        if not command:
            return jsonify({
                'success': False,
                'message': '命令不能为空'
            }), 400
        
        # 执行命令
        success, message, cmd_record = webshell_terminal_service.execute_command(
            session_id, command, timeout
        )
        
        return jsonify({
            'success': success,
            'data': {
                'session_id': session_id,
                'command_record': cmd_record,
                'executed_at': datetime.utcnow().isoformat()
            },
            'message': message
        })
        
    except Exception as e:
        logger.error(f"Execute WebShell command error: {e}")
        return jsonify({
            'success': False,
            'message': '执行命令失败'
        }), 500


@hosts_bp.route('/webshell/<session_id>/resize', methods=['POST'])
@tenant_required
def resize_webshell_terminal(session_id):
    """调整 WebShell 终端大小"""
    try:
        from app.services.webshell_service import webshell_service
        from app.services.webshell_terminal_service import webshell_terminal_service
        
        session_info = webshell_service.get_session(session_id)
        
        if not session_info:
            return jsonify({
                'success': False,
                'message': 'WebShell 会话不存在'
            }), 404
        
        # 验证会话所有权
        if session_info['user_id'] != g.user_id:
            return jsonify({
                'success': False,
                'message': '无权访问该会话'
            }), 403
        
        data = request.get_json()
        cols = data.get('cols', 80)
        rows = data.get('rows', 24)
        
        # 验证参数
        if not isinstance(cols, int) or not isinstance(rows, int):
            return jsonify({
                'success': False,
                'message': '终端大小参数必须是整数'
            }), 400
        
        if cols < 10 or cols > 300 or rows < 5 or rows > 100:
            return jsonify({
                'success': False,
                'message': '终端大小参数超出有效范围'
            }), 400
        
        # 调整终端大小
        success = webshell_terminal_service.resize_terminal(session_id, cols, rows)
        
        if success:
            return jsonify({
                'success': True,
                'data': {
                    'session_id': session_id,
                    'terminal_size': {
                        'cols': cols,
                        'rows': rows
                    }
                },
                'message': '终端大小调整成功'
            })
        else:
            return jsonify({
                'success': False,
                'message': '调整终端大小失败'
            }), 500
        
    except Exception as e:
        logger.error(f"Resize WebShell terminal error: {e}")
        return jsonify({
            'success': False,
            'message': '调整终端大小失败'
        }), 500


@hosts_bp.route('/webshell/<session_id>/terminate', methods=['POST'])
@tenant_required
def terminate_webshell_session(session_id):
    """终止 WebShell 会话"""
    try:
        from app.services.webshell_service import webshell_service
        from app.services.webshell_terminal_service import webshell_terminal_service
        
        session_info = webshell_service.get_session(session_id)
        
        if not session_info:
            return jsonify({
                'success': False,
                'message': 'WebShell 会话不存在'
            }), 404
        
        # 验证会话所有权
        if session_info['user_id'] != g.user_id:
            return jsonify({
                'success': False,
                'message': '无权访问该会话'
            }), 403
        
        # 终止终端会话
        webshell_terminal_service.terminate_terminal_session(session_id)
        
        # 终止 WebShell 会话
        webshell_service.terminate_session(session_id)
        
        return jsonify({
            'success': True,
            'data': {
                'session_id': session_id,
                'terminated_at': datetime.utcnow().isoformat()
            },
            'message': 'WebShell 会话已终止'
        })
        
    except Exception as e:
        logger.error(f"Terminate WebShell session error: {e}")
        return jsonify({
            'success': False,
            'message': '终止 WebShell 会话失败'
        }), 500


@hosts_bp.route('/webshell/sessions', methods=['GET'])
@tenant_required
def list_webshell_sessions():
    """获取用户的 WebShell 会话列表"""
    try:
        from app.services.webshell_service import webshell_service
        from app.services.webshell_terminal_service import webshell_terminal_service
        
        # 获取用户的所有会话
        user_sessions = webshell_service.get_user_sessions(g.user_id)
        
        sessions_data = []
        for session_info in user_sessions:
            # 获取主机信息
            host = SSHHost.query.filter_by(
                id=session_info['host_id'],
                tenant_id=g.tenant_id
            ).first()
            
            if not host:
                continue
            
            session_id = session_info['session_id']
            
            # 获取终端会话状态
            terminal_session = webshell_terminal_service.get_terminal_session(session_id)
            terminal_active = terminal_session.is_active if terminal_session else False
            
            sessions_data.append({
                'session_id': session_id,
                'host_id': host.id,
                'host_name': host.name,
                'hostname': host.hostname,
                'username': host.username,
                'webshell_status': session_info['status'],
                'terminal_active': terminal_active,
                'terminal_size': session_info.get('terminal_size', {'cols': 80, 'rows': 24}),
                'created_at': session_info['created_at'],
                'last_activity': session_info.get('last_activity'),
                'is_connected': session_info['status'] == 'active' and terminal_active
            })
        
        return jsonify({
            'success': True,
            'data': {
                'sessions': sessions_data,
                'total_count': len(sessions_data)
            }
        })
        
    except Exception as e:
        logger.error(f"List WebShell sessions error: {e}")
        return jsonify({
            'success': False,
            'message': '获取 WebShell 会话列表失败'
        }), 500



# ============================================
# 主机探测 API 端点
# ============================================

@hosts_bp.route('/<int:host_id>/probe', methods=['POST'])
@role_required('admin', 'super_admin', '超级管理员', '运维管理员', '系统管理员')
def probe_host(host_id):
    """
    探测单个主机
    
    使用 Ansible ping 模块检测主机连通性，任务异步执行
    
    Requirements: 6.1, 8.2
    """
    try:
        from app.tasks.host_probe_tasks import execute_ansible_ping
        from celery.exceptions import OperationalError
        
        # 验证主机存在
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        # 获取超时参数（支持无请求体的情况）
        data = request.get_json(silent=True) or {}
        timeout = data.get('timeout', 30)
        
        # 验证超时参数
        if not isinstance(timeout, int) or timeout < 5 or timeout > 300:
            return jsonify({
                'success': False,
                'message': '超时时间必须在 5-300 秒之间'
            }), 400
        
        try:
            # 提交 Celery 任务
            task = execute_ansible_ping.apply_async(
                kwargs={
                    'host_id': host_id,
                    'tenant_id': g.tenant_id,
                    'timeout': timeout
                }
            )
            
            # 更新主机探测状态为 pending
            host.last_probe_status = 'pending'
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'task_id': task.id,
                    'host_id': host_id,
                    'host_name': host.name,
                    'hostname': host.hostname,
                    'status': 'pending',
                    'message': '探测任务已提交'
                },
                'message': '探测任务已提交'
            })
            
        except OperationalError as e:
            logger.error(f"Celery 服务不可用: {e}")
            return jsonify({
                'success': False,
                'message': '探测服务暂时不可用，请稍后重试'
            }), 503
            
    except Exception as e:
        db.session.rollback()
        logger.error(f"Probe host error: {e}")
        return jsonify({
            'success': False,
            'message': '探测任务提交失败'
        }), 500


@hosts_bp.route('/probe/batch', methods=['POST'])
@role_required('admin', 'super_admin', '超级管理员', '运维管理员', '系统管理员')
def probe_hosts_batch():
    """
    批量探测主机
    
    对多个主机执行 Ansible ping 探测，任务异步执行
    
    Requirements: 6.2, 6.7
    """
    try:
        from app.tasks.host_probe_tasks import execute_ansible_ping
        from celery.exceptions import OperationalError
        
        data = request.get_json() or {}
        host_ids = data.get('host_ids', [])
        timeout = data.get('timeout', 30)
        
        # 验证参数
        if not host_ids or not isinstance(host_ids, list):
            return jsonify({
                'success': False,
                'message': '请提供有效的主机 ID 列表'
            }), 400
        
        if len(host_ids) > 100:
            return jsonify({
                'success': False,
                'message': '单次批量探测最多支持 100 台主机'
            }), 400
        
        # 验证超时参数
        if not isinstance(timeout, int) or timeout < 5 or timeout > 300:
            return jsonify({
                'success': False,
                'message': '超时时间必须在 5-300 秒之间'
            }), 400
        
        # 验证所有主机存在且属于当前租户
        hosts = SSHHost.query.filter(
            SSHHost.id.in_(host_ids),
            SSHHost.tenant_id == g.tenant_id
        ).all()
        
        if len(hosts) != len(host_ids):
            found_ids = {host.id for host in hosts}
            missing_ids = [hid for hid in host_ids if hid not in found_ids]
            return jsonify({
                'success': False,
                'message': f'部分主机不存在或无权访问: {missing_ids}'
            }), 404
        
        try:
            results = []
            for host in hosts:
                # 提交 Celery 任务
                task = execute_ansible_ping.apply_async(
                    kwargs={
                        'host_id': host.id,
                        'tenant_id': g.tenant_id,
                        'timeout': timeout
                    }
                )
                
                # 更新主机探测状态为 pending
                host.last_probe_status = 'pending'
                
                results.append({
                    'host_id': host.id,
                    'host_name': host.name,
                    'hostname': host.hostname,
                    'task_id': task.id,
                    'status': 'pending'
                })
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'total': len(results),
                    'queued': len(results),
                    'results': results
                },
                'message': f'已提交 {len(results)} 个探测任务'
            })
            
        except OperationalError as e:
            logger.error(f"Celery 服务不可用: {e}")
            return jsonify({
                'success': False,
                'message': '探测服务暂时不可用，请稍后重试'
            }), 503
            
    except Exception as e:
        db.session.rollback()
        logger.error(f"Batch probe hosts error: {e}")
        return jsonify({
            'success': False,
            'message': '批量探测任务提交失败'
        }), 500


@hosts_bp.route('/probe/group/<int:group_id>', methods=['POST'])
@role_required('admin', 'super_admin', '超级管理员', '运维管理员', '系统管理员')
def probe_group(group_id):
    """
    探测分组内所有主机
    
    对指定分组内的所有主机执行 Ansible ping 探测
    
    Requirements: 6.2, 6.7
    """
    try:
        from app.models.host import HostGroup
        from app.tasks.host_probe_tasks import execute_ansible_ping
        from celery.exceptions import OperationalError
        
        # 验证分组存在
        group = HostGroup.query.filter_by(
            id=group_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not group:
            return jsonify({
                'success': False,
                'message': '分组不存在'
            }), 404
        
        # 获取超时参数
        data = request.get_json() or {}
        timeout = data.get('timeout', 30)
        
        # 验证超时参数
        if not isinstance(timeout, int) or timeout < 5 or timeout > 300:
            return jsonify({
                'success': False,
                'message': '超时时间必须在 5-300 秒之间'
            }), 400
        
        # 获取分组内所有主机
        hosts = SSHHost.query.filter_by(
            group_id=group_id,
            tenant_id=g.tenant_id
        ).all()
        
        if not hosts:
            return jsonify({
                'success': True,
                'data': {
                    'group_id': group_id,
                    'group_name': group.name,
                    'total': 0,
                    'queued': 0,
                    'results': []
                },
                'message': '分组内没有主机'
            })
        
        try:
            results = []
            for host in hosts:
                # 提交 Celery 任务
                task = execute_ansible_ping.apply_async(
                    kwargs={
                        'host_id': host.id,
                        'tenant_id': g.tenant_id,
                        'timeout': timeout
                    }
                )
                
                # 更新主机探测状态为 pending
                host.last_probe_status = 'pending'
                
                results.append({
                    'host_id': host.id,
                    'host_name': host.name,
                    'hostname': host.hostname,
                    'task_id': task.id,
                    'status': 'pending'
                })
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'data': {
                    'group_id': group_id,
                    'group_name': group.name,
                    'total': len(results),
                    'queued': len(results),
                    'results': results
                },
                'message': f'已提交 {len(results)} 个探测任务'
            })
            
        except OperationalError as e:
            logger.error(f"Celery 服务不可用: {e}")
            return jsonify({
                'success': False,
                'message': '探测服务暂时不可用，请稍后重试'
            }), 503
            
    except Exception as e:
        db.session.rollback()
        logger.error(f"Probe group error: {e}")
        return jsonify({
            'success': False,
            'message': '分组探测任务提交失败'
        }), 500


@hosts_bp.route('/probe/task/<task_id>', methods=['GET'])
@tenant_required
def get_probe_task_status(task_id):
    """
    获取探测任务状态
    
    查询 Celery 任务的执行状态和结果
    
    Requirements: 8.3
    """
    try:
        from app.celery_app import celery
        
        # 获取任务状态
        task = celery.AsyncResult(task_id)
        
        # 构建响应数据
        celery_status = task.status.lower() if task.status else 'unknown'
        
        # 映射 Celery 状态到前端友好的状态
        status_mapping = {
            'pending': 'pending',
            'started': 'running',
            'success': 'success',
            'failure': 'failed',
            'retry': 'running',
            'revoked': 'failed'
        }
        
        response_data = {
            'task_id': task_id,
            'status': status_mapping.get(celery_status, celery_status)
        }
        
        # 如果任务完成，获取结果
        if task.ready():
            if task.successful():
                result = task.result
                if isinstance(result, dict):
                    # 获取实际探测结果
                    probe_status = result.get('status', 'unknown')
                    probe_message = result.get('message', '')
                    
                    # 使用实际探测结果状态作为外层状态
                    response_data['status'] = probe_status
                    response_data['result'] = {
                        'host_id': result.get('host_id'),
                        'status': probe_status,
                        'message': probe_message,
                        'response_time': result.get('response_time'),
                        'ansible_output': result.get('ansible_output')
                    }
                    
                    # 如果探测失败，确保错误信息传递
                    if probe_status == 'failed':
                        response_data['error'] = probe_message
            else:
                # Celery 任务本身失败（异常）
                response_data['status'] = 'failed'
                error_msg = str(task.result) if task.result else '任务执行失败'
                # 解析常见错误
                if 'ConnectionRefusedError' in error_msg or 'Connection refused' in error_msg:
                    error_msg = 'Celery 连接失败，请检查 Redis 服务是否正常'
                elif 'TimeoutError' in error_msg or 'timed out' in error_msg.lower():
                    error_msg = '任务执行超时'
                response_data['error'] = error_msg
                response_data['result'] = {
                    'status': 'failed',
                    'message': error_msg
                }
        
        return jsonify({
            'success': True,
            'data': response_data
        })
        
    except Exception as e:
        logger.error(f"Get probe task status error: {e}")
        return jsonify({
            'success': False,
            'message': f'获取任务状态失败: {str(e)}'
        }), 500


@hosts_bp.route('/<int:host_id>/probe-history', methods=['GET'])
@tenant_required
def get_probe_history(host_id):
    """
    获取主机探测历史记录
    
    查询指定主机的探测历史
    
    Requirements: 7.2
    """
    try:
        from app.models.host import HostProbeResult
        
        # 验证主机存在
        host = SSHHost.query.filter_by(
            id=host_id,
            tenant_id=g.tenant_id
        ).first()
        
        if not host:
            return jsonify({
                'success': False,
                'message': '主机不存在'
            }), 404
        
        # 获取查询参数
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        limit = request.args.get('limit', type=int)
        
        # 限制每页最大数量
        per_page = min(per_page, 100)
        
        # 构建查询
        query = HostProbeResult.query.filter_by(host_id=host_id).order_by(
            HostProbeResult.probed_at.desc()
        )
        
        # 如果指定了 limit，直接返回最近的记录
        if limit:
            limit = min(limit, 1000)
            results = query.limit(limit).all()
            
            return jsonify({
                'success': True,
                'data': {
                    'host_id': host_id,
                    'host_name': host.name,
                    'hostname': host.hostname,
                    'history': [r.to_dict() for r in results],
                    'total': len(results)
                }
            })
        
        # 分页查询
        pagination = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        return jsonify({
            'success': True,
            'data': {
                'host_id': host_id,
                'host_name': host.name,
                'hostname': host.hostname,
                'history': [r.to_dict() for r in pagination.items],
                'pagination': {
                    'page': pagination.page,
                    'per_page': pagination.per_page,
                    'total': pagination.total,
                    'pages': pagination.pages,
                    'has_prev': pagination.has_prev,
                    'has_next': pagination.has_next
                }
            }
        })
        
    except Exception as e:
        logger.error(f"Get probe history error: {e}")
        return jsonify({
            'success': False,
            'message': '获取探测历史失败'
        }), 500


# ============================================
# 主机批量导入 API 端点
# ============================================

@hosts_bp.route('/import', methods=['POST'])
@role_required('admin', 'super_admin', '超级管理员', '运维管理员', '系统管理员')
def import_hosts():
    """
    批量导入主机（Excel 文件）
    
    Excel 格式要求：
    - 主机名称（必填）
    - 主机地址（必填）
    - 端口（默认22）
    - 用户名（必填）
    - 密码认证（T/F）
    - 密钥认证（T/F）
    - 密码
    - 私钥
    - 分组名称
    - 操作系统
    - 描述
    """
    try:
        from openpyxl import load_workbook
        from io import BytesIO
        from app.models.host import HostGroup
        
        # 检查是否有文件上传
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '请上传 Excel 文件'
            }), 400
        
        file = request.files['file']
        
        if not file.filename:
            return jsonify({
                'success': False,
                'message': '请选择文件'
            }), 400
        
        # 检查文件扩展名
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'message': '仅支持 .xlsx 或 .xls 格式的 Excel 文件'
            }), 400
        
        # 读取 Excel 文件
        try:
            workbook = load_workbook(BytesIO(file.read()))
            sheet = workbook.active
        except Exception as e:
            logger.error(f"读取 Excel 文件失败: {e}")
            return jsonify({
                'success': False,
                'message': '无法读取 Excel 文件，请确保文件格式正确'
            }), 400
        
        # 获取表头（第一行）
        headers = [cell.value for cell in sheet[1]]
        
        # 预期的列名映射
        expected_columns = {
            '主机名称': 'name',
            '主机地址': 'hostname',
            '端口': 'port',
            '用户名': 'username',
            '密码认证': 'use_password',
            '密钥认证': 'use_key',
            '密码': 'password',
            '私钥': 'private_key',
            '分组名称': 'group_name',
            '操作系统': 'os_type',
            '描述': 'description'
        }
        
        # 建立列索引映射
        column_map = {}
        for idx, header in enumerate(headers):
            if header in expected_columns:
                column_map[expected_columns[header]] = idx
        
        # 验证必需列
        required_columns = ['name', 'hostname', 'username']
        missing_columns = [col for col in required_columns if col not in column_map]
        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'Excel 缺少必需列: {", ".join([k for k, v in expected_columns.items() if v in missing_columns])}'
            }), 400
        
        # 预加载分组信息
        groups = HostGroup.query.filter_by(tenant_id=g.tenant_id).all()
        group_map = {group.name: group.id for group in groups}
        
        # 解析数据行
        results = {
            'success': [],
            'failed': [],
            'skipped': []
        }
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 跳过空行
                if not any(row):
                    continue
                
                # 获取字段值
                def get_value(field, default=None):
                    if field in column_map:
                        val = row[column_map[field]]
                        return val if val is not None else default
                    return default
                
                name = get_value('name', '').strip() if get_value('name') else ''
                hostname = get_value('hostname', '').strip() if get_value('hostname') else ''
                username = get_value('username', '').strip() if get_value('username') else ''
                
                # 验证必填字段
                if not name or not hostname or not username:
                    results['failed'].append({
                        'row': row_idx,
                        'name': name or '(空)',
                        'hostname': hostname or '(空)',
                        'error': '主机名称、主机地址、用户名为必填项'
                    })
                    continue
                
                # 解析端口
                port_val = get_value('port', 22)
                try:
                    port = int(port_val) if port_val else 22
                    if port < 1 or port > 65535:
                        port = 22
                except (ValueError, TypeError):
                    port = 22
                
                # 解析认证方式（T/F 格式）
                use_password_val = str(get_value('use_password', '')).strip().upper()
                use_key_val = str(get_value('use_key', '')).strip().upper()
                
                use_password = use_password_val in ('T', 'TRUE', '是', '1', 'Y', 'YES')
                use_key = use_key_val in ('T', 'TRUE', '是', '1', 'Y', 'YES')
                
                # 确定认证类型
                if use_password and not use_key:
                    auth_type = 'password'
                elif use_key and not use_password:
                    auth_type = 'key'
                elif use_password and use_key:
                    # 两个都选了，优先密码
                    auth_type = 'password'
                else:
                    # 都没选，默认密码
                    auth_type = 'password'
                
                # 获取认证凭据
                password = get_value('password', '')
                private_key = get_value('private_key', '')
                
                # 验证认证信息
                if auth_type == 'password' and not password:
                    results['failed'].append({
                        'row': row_idx,
                        'name': name,
                        'hostname': hostname,
                        'error': '选择密码认证但未提供密码'
                    })
                    continue
                
                if auth_type == 'key' and not private_key:
                    results['failed'].append({
                        'row': row_idx,
                        'name': name,
                        'hostname': hostname,
                        'error': '选择密钥认证但未提供私钥'
                    })
                    continue
                
                # 检查主机是否已存在
                existing_host = SSHHost.query.filter_by(
                    tenant_id=g.tenant_id,
                    hostname=hostname,
                    port=port
                ).first()
                
                if existing_host:
                    results['skipped'].append({
                        'row': row_idx,
                        'name': name,
                        'hostname': hostname,
                        'reason': f'主机已存在 (ID: {existing_host.id})'
                    })
                    continue
                
                # 处理分组
                group_name = get_value('group_name', '').strip() if get_value('group_name') else ''
                group_id = None
                if group_name:
                    if group_name in group_map:
                        group_id = group_map[group_name]
                    else:
                        # 自动创建分组
                        new_group = HostGroup(
                            tenant_id=g.tenant_id,
                            name=group_name,
                            description=f'通过批量导入自动创建'
                        )
                        db.session.add(new_group)
                        db.session.flush()  # 获取 ID
                        group_map[group_name] = new_group.id
                        group_id = new_group.id
                
                # 创建主机
                host = SSHHost(
                    tenant_id=g.tenant_id,
                    name=name,
                    hostname=hostname,
                    port=port,
                    username=username,
                    auth_type=auth_type,
                    password=password if auth_type == 'password' else None,
                    private_key=private_key if auth_type == 'key' else None,
                    group_id=group_id,
                    os_type=get_value('os_type', '').strip() if get_value('os_type') else '',
                    description=get_value('description', '').strip() if get_value('description') else '',
                    status=1
                )
                
                db.session.add(host)
                db.session.flush()  # 获取 ID
                
                results['success'].append({
                    'row': row_idx,
                    'id': host.id,
                    'name': name,
                    'hostname': hostname
                })
                
            except Exception as e:
                logger.error(f"处理第 {row_idx} 行时出错: {e}")
                results['failed'].append({
                    'row': row_idx,
                    'name': get_value('name', '(未知)') if 'get_value' in dir() else '(未知)',
                    'hostname': get_value('hostname', '(未知)') if 'get_value' in dir() else '(未知)',
                    'error': str(e)
                })
        
        # 提交事务
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': {
                'total': len(results['success']) + len(results['failed']) + len(results['skipped']),
                'success_count': len(results['success']),
                'failed_count': len(results['failed']),
                'skipped_count': len(results['skipped']),
                'success': results['success'],
                'failed': results['failed'],
                'skipped': results['skipped']
            },
            'message': f"导入完成：成功 {len(results['success'])} 条，失败 {len(results['failed'])} 条，跳过 {len(results['skipped'])} 条"
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Import hosts error: {e}")
        return jsonify({
            'success': False,
            'message': f'导入失败: {str(e)}'
        }), 500


@hosts_bp.route('/import/template', methods=['GET'])
@tenant_required
def download_import_template():
    """
    下载主机导入模板
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "主机导入模板"
        
        # 定义表头
        headers = [
            '主机名称', '主机地址', '端口', '用户名',
            '密码认证', '密钥认证', '密码', '私钥',
            '分组名称', '操作系统', '描述'
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 添加示例数据
        example_data = [
            ['Web服务器01', '192.168.1.100', 22, 'root', 'T', 'F', 'password123', '', '生产环境', 'CentOS 7', 'Web应用服务器'],
            ['数据库服务器', '192.168.1.101', 22, 'admin', 'F', 'T', '', '-----BEGIN RSA PRIVATE KEY-----\n...\n-----END RSA PRIVATE KEY-----', '生产环境', 'Ubuntu 20.04', 'MySQL数据库'],
            ['测试服务器', '10.0.0.50', 2222, 'test', 'T', 'F', 'test123', '', '测试环境', 'Debian 11', '测试用服务器'],
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # 设置列宽
        column_widths = [15, 18, 8, 12, 12, 12, 15, 30, 15, 15, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # 添加说明行
        ws.cell(row=6, column=1, value='说明：')
        ws.cell(row=6, column=1).font = Font(bold=True)
        
        instructions = [
            '1. 主机名称、主机地址、用户名为必填项',
            '2. 端口默认为 22，可不填',
            '3. 密码认证和密钥认证列使用 T 或 F 表示选择（T=选择，F=不选择）',
            '4. 如果选择密码认证（T），则密码列必填；如果选择密钥认证（T），则私钥列必填',
            '5. 如果两个认证都选择 T，将优先使用密码认证',
            '6. 分组名称如果不存在会自动创建',
            '7. 已存在的主机（相同地址和端口）会被跳过'
        ]
        
        for idx, instruction in enumerate(instructions, 7):
            ws.cell(row=idx, column=1, value=instruction)
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=11)
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='主机导入模板.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Download import template error: {e}")
        return jsonify({
            'success': False,
            'message': '下载模板失败'
        }), 500


@hosts_bp.route('/stats', methods=['GET'])
@tenant_required
def get_host_stats():
    """
    获取主机统计信息（用于仪表盘）
    
    Returns:
        JSON response with host statistics:
        - online: 在线主机数量
        - offline: 离线主机数量
        - total: 总主机数量
    """
    try:
        # 获取当前租户的所有主机
        hosts = SSHHost.query.filter_by(tenant_id=g.tenant_id).all()
        
        stats = {
            'online': 0,
            'offline': 0,
            'total': len(hosts)
        }
        
        # 统计各状态的主机数量
        # status字段: 1=活跃, 0=禁用
        for host in hosts:
            if host.status == 1:
                stats['online'] += 1
            else:
                stats['offline'] += 1
        
        return jsonify({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        logger.error(f"Get host stats error: {e}")
        return jsonify({
            'success': False,
            'message': '获取主机统计失败',
            'error': str(e)
        }), 500
